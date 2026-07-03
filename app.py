"""
GSTR-1 HSN Summary Calculator  ·  v2
Extract HSN-wise summaries from invoices (images/PDF via Claude Vision, Excel/CSV parsed free).
Run: python app.py
"""

import os, json, base64, threading, re, io, csv
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import customtkinter as ctk
import anthropic
import PIL.Image
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF
from pathlib import Path

# ──────────────────────────────  Design tokens  ──────────────────────────────
ACCENT        = "#4F6EF7"   # primary indigo
ACCENT_HOVER  = "#3D5AE0"
SUCCESS       = "#2ECC71"
WARNING       = "#F5A623"
DANGER        = "#E74C3C"
MUTED         = ("gray45", "gray60")
CARD_BG       = ("#F4F6FB", "#23262E")
SIDEBAR_BG    = ("#EDF0F7", "#1B1D23")
FONT_FAMILY   = "Segoe UI"

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

CONFIG_FILE = Path.home() / ".gstr1_config.json"

HSN_ALIASES  = ["hsn", "hsn code", "hsn/sac", "hsncode", "hsnsac", "hsn_code", "sac", "sac code"]
QTY_ALIASES  = ["quantity", "qty", "units", "nos"]
UQC_ALIASES  = ["uqc", "unit", "uom", "u/m", "unit of measurement"]
TAX_ALIASES  = ["taxable value", "taxable amount", "taxable", "taxable val",
                 "tax value", "assessable value", "basic amount", "taxable amt"]
IGST_ALIASES = ["igst", "igst amount", "integrated tax", "igst amt"]
CGST_ALIASES = ["cgst", "cgst amount", "central tax", "cgst amt"]
SGST_ALIASES = ["sgst", "sgst amount", "state tax", "sgst amt", "utgst"]

EXTRACT_PROMPT = """You are a GST invoice data extractor for Indian invoices.
Analyze this invoice carefully and extract ALL line items.

For each line item return:
- hsn: HSN/SAC code exactly as printed (string). If absent, use "UNKNOWN"
- uqc: Unit of quantity (NOS, KGS, MTR, LTR, PCS, BAG, BOX, etc.)
- quantity: numeric quantity (number)
- taxable_value: taxable/assessable amount in INR (number, no commas)
- igst: IGST amount (number, 0 if absent)
- cgst: CGST amount (number, 0 if absent)
- sgst: SGST/UTGST amount (number, 0 if absent)

Rules:
- Return ONLY a valid JSON array, no explanation, no markdown fences
- All monetary values as plain numbers (no commas, no currency symbols)
- If a tax column is missing, use 0
- Do NOT skip any line item
- If HSN is truly absent, use "UNKNOWN"

Example: [{"hsn":"1234","uqc":"NOS","quantity":10,"taxable_value":5000,"igst":900,"cgst":0,"sgst":0}]"""


def match_col(headers, aliases):
    for a in aliases:
        for i, h in enumerate(headers):
            hn = str(h or "").lower().strip()
            if hn == a or a in hn:
                return i
    return -1

def to_num(v):
    try:
        return float(re.sub(r"[^\d.\-]", "", str(v or "")))
    except Exception:
        return 0.0


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("GSTR-1 HSN Calculator")
        self.geometry("1180x780")
        self.minsize(980, 640)
        self.files: list = []
        self.hsn_map: dict = {}
        self._load_config()
        self._build_ui()

    # ─────────────────────────────  Config  ─────────────────────────────
    def _load_config(self):
        self.api_key = ""
        if CONFIG_FILE.exists():
            try:
                data = json.loads(CONFIG_FILE.read_text())
                self.api_key = data.get("api_key", "")
            except Exception:
                pass

    def _save_config(self):
        CONFIG_FILE.write_text(json.dumps({"api_key": self.api_key}))

    # ─────────────────────────────  Fonts  ─────────────────────────────
    def _font(self, size=12, weight="normal"):
        return ctk.CTkFont(family=FONT_FAMILY, size=size, weight=weight)

    # ─────────────────────────────  UI shell  ─────────────────────────────
    def _build_ui(self):
        self.sidebar = ctk.CTkFrame(self, width=280, corner_radius=0, fg_color=SIDEBAR_BG)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar()

        self.main = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.main.pack(side="right", fill="both", expand=True)

        self.tabview = ctk.CTkTabview(
            self.main,
            segmented_button_selected_color=ACCENT,
            segmented_button_selected_hover_color=ACCENT_HOVER)
        self.tabview.pack(fill="both", expand=True, padx=16, pady=(8, 16))
        self.tab_upload  = self.tabview.add("  1 · Upload  ")
        self.tab_process = self.tabview.add("  2 · Process  ")
        self.tab_results = self.tabview.add("  3 · Review & Export  ")
        self.tabview.set("  1 · Upload  ")
        self._build_upload_tab()
        self._build_process_tab()
        self._build_results_tab()

    # ─────────────────────────────  Sidebar  ─────────────────────────────
    def _build_sidebar(self):
        sb = self.sidebar

        # Brand block
        brand = ctk.CTkFrame(sb, fg_color="transparent")
        brand.pack(fill="x", padx=20, pady=(26, 18))
        ctk.CTkLabel(brand, text="GSTR-1", font=self._font(24, "bold"),
                     text_color=ACCENT).pack(anchor="w")
        ctk.CTkLabel(brand, text="HSN Summary Calculator",
                     font=self._font(13)).pack(anchor="w")
        ctk.CTkLabel(brand, text="v2.0  ·  Claude Vision", text_color=MUTED,
                     font=self._font(10)).pack(anchor="w", pady=(2, 0))

        # API key card
        key_card = ctk.CTkFrame(sb, corner_radius=10, fg_color=CARD_BG)
        key_card.pack(fill="x", padx=14, pady=(0, 14))
        ctk.CTkLabel(key_card, text="ANTHROPIC API KEY", font=self._font(10, "bold"),
                     text_color=MUTED, anchor="w").pack(fill="x", padx=14, pady=(12, 4))

        key_row = ctk.CTkFrame(key_card, fg_color="transparent")
        key_row.pack(fill="x", padx=14)
        self.api_entry = ctk.CTkEntry(key_row, show="•", placeholder_text="sk-ant-...",
                                       font=self._font(11), border_color=ACCENT)
        self.api_entry.pack(side="left", fill="x", expand=True)
        if self.api_key:
            self.api_entry.insert(0, self.api_key)
        self._key_visible = False
        self.eye_btn = ctk.CTkButton(key_row, text="👁", width=34, fg_color="transparent",
                                      border_width=1, text_color=MUTED,
                                      command=self._toggle_key_visibility)
        self.eye_btn.pack(side="left", padx=(6, 0))

        ctk.CTkButton(key_card, text="Save Key", height=30, fg_color=ACCENT,
                      hover_color=ACCENT_HOVER, font=self._font(12, "bold"),
                      command=self._save_key).pack(fill="x", padx=14, pady=(8, 4))
        self.key_status = ctk.CTkLabel(
            key_card,
            text="●  Key saved" if self.api_key else "○  No key saved",
            text_color=SUCCESS if self.api_key else MUTED,
            font=self._font(10), anchor="w")
        self.key_status.pack(fill="x", padx=14, pady=(0, 4))
        ctk.CTkLabel(key_card, text="console.anthropic.com", text_color=MUTED,
                     font=self._font(9), anchor="w").pack(fill="x", padx=14, pady=(0, 12))

        # Formats card
        fmt_card = ctk.CTkFrame(sb, corner_radius=10, fg_color=CARD_BG)
        fmt_card.pack(fill="x", padx=14, pady=(0, 14))
        ctk.CTkLabel(fmt_card, text="SUPPORTED FORMATS", font=self._font(10, "bold"),
                     text_color=MUTED, anchor="w").pack(fill="x", padx=14, pady=(12, 4))
        for icon, txt, note in [
            ("🖼", "JPG / PNG", "Claude Vision"),
            ("📄", "PDF", "pages → images → Vision"),
            ("📊", "Excel / CSV", "parsed directly · free"),
        ]:
            row = ctk.CTkFrame(fmt_card, fg_color="transparent")
            row.pack(fill="x", padx=14, pady=1)
            ctk.CTkLabel(row, text=f"{icon}  {txt}", font=self._font(11),
                         anchor="w").pack(side="left")
            ctk.CTkLabel(row, text=note, font=self._font(9), text_color=MUTED,
                         anchor="e").pack(side="right")
        ctk.CTkLabel(fmt_card, text="", font=self._font(2)).pack(pady=(0, 6))

        # Files card
        files_card = ctk.CTkFrame(sb, corner_radius=10, fg_color=CARD_BG)
        files_card.pack(fill="both", expand=True, padx=14, pady=(0, 10))
        head = ctk.CTkFrame(files_card, fg_color="transparent")
        head.pack(fill="x", padx=14, pady=(12, 4))
        ctk.CTkLabel(head, text="QUEUE", font=self._font(10, "bold"),
                     text_color=MUTED, anchor="w").pack(side="left")
        self.file_count_chip = ctk.CTkLabel(head, text="0", width=28, corner_radius=8,
                                             fg_color=ACCENT, text_color="white",
                                             font=self._font(10, "bold"))
        self.file_count_chip.pack(side="right")
        self.file_listbox_frame = ctk.CTkScrollableFrame(files_card, fg_color="transparent")
        self.file_listbox_frame.pack(fill="both", expand=True, padx=8)
        ctk.CTkButton(files_card, text="Clear all", height=28, fg_color="transparent",
                      border_width=1, text_color=MUTED, font=self._font(11),
                      command=self._clear_files).pack(fill="x", padx=14, pady=(4, 12))

        # Theme toggle
        self.theme_switch = ctk.CTkSwitch(sb, text="Light mode", font=self._font(11),
                                           progress_color=ACCENT, command=self._toggle_theme)
        self.theme_switch.pack(padx=20, pady=(0, 16), anchor="w")

    def _toggle_key_visibility(self):
        self._key_visible = not self._key_visible
        self.api_entry.configure(show="" if self._key_visible else "•")

    def _toggle_theme(self):
        mode = "light" if self.theme_switch.get() else "dark"
        ctk.set_appearance_mode(mode)
        self._style_tree()

    # ─────────────────────────────  Upload tab  ─────────────────────────────
    def _build_upload_tab(self):
        tab = self.tab_upload

        self.drop_frame = ctk.CTkFrame(tab, height=230, corner_radius=14,
                                        border_width=2, border_color=ACCENT,
                                        fg_color=CARD_BG)
        self.drop_frame.pack(fill="x", padx=24, pady=(24, 14))
        self.drop_frame.pack_propagate(False)
        ctk.CTkLabel(self.drop_frame, text="📂", font=ctk.CTkFont(size=44)).pack(pady=(34, 6))
        ctk.CTkLabel(self.drop_frame, text="Select your invoices",
                     font=self._font(17, "bold")).pack()
        ctk.CTkLabel(self.drop_frame, text="JPG · PNG · PDF · XLSX · XLS · CSV",
                     text_color=MUTED, font=self._font(11)).pack(pady=(2, 0))
        ctk.CTkButton(self.drop_frame, text="Browse Files", width=180, height=38,
                      fg_color=ACCENT, hover_color=ACCENT_HOVER,
                      font=self._font(13, "bold"),
                      command=self._browse_files).pack(pady=(14, 0))

        self.upload_status = ctk.CTkLabel(tab, text="No files selected.",
                                           text_color=MUTED, font=self._font(12))
        self.upload_status.pack(pady=(0, 6))

        self.process_btn = ctk.CTkButton(
            tab, text="⚡  Process Bills", font=self._font(15, "bold"),
            height=52, corner_radius=12, fg_color=ACCENT, hover_color=ACCENT_HOVER,
            state="disabled", command=self._start_processing)
        self.process_btn.pack(fill="x", padx=24, pady=(4, 8))

        ctk.CTkLabel(tab, text="Images & PDFs are read by Claude Haiku Vision (uses API credits). "
                               "Excel/CSV files are parsed locally at no cost.",
                     text_color=MUTED, font=self._font(10), wraplength=760).pack(padx=24)

    # ─────────────────────────────  Process tab  ─────────────────────────────
    def _build_process_tab(self):
        tab = self.tab_process

        top = ctk.CTkFrame(tab, fg_color="transparent")
        top.pack(fill="x", padx=24, pady=(20, 6))
        ctk.CTkLabel(top, text="Processing", font=self._font(15, "bold"),
                     anchor="w").pack(side="left")
        self.progress_pct = ctk.CTkLabel(top, text="0%", font=self._font(13, "bold"),
                                          text_color=ACCENT)
        self.progress_pct.pack(side="right")

        self.progress_bar = ctk.CTkProgressBar(tab, height=10, progress_color=ACCENT)
        self.progress_bar.pack(fill="x", padx=24, pady=(0, 6))
        self.progress_bar.set(0)
        self.progress_label = ctk.CTkLabel(tab, text="Waiting...", text_color=MUTED,
                                            font=self._font(11), anchor="w")
        self.progress_label.pack(fill="x", padx=24, pady=(0, 8))

        log_card = ctk.CTkFrame(tab, corner_radius=12, fg_color=CARD_BG)
        log_card.pack(fill="both", expand=True, padx=24, pady=(0, 20))
        ctk.CTkLabel(log_card, text="LOG", font=self._font(10, "bold"),
                     text_color=MUTED, anchor="w").pack(fill="x", padx=14, pady=(10, 0))
        self.log_box = ctk.CTkTextbox(log_card, font=ctk.CTkFont(family="Consolas", size=11),
                                       fg_color="transparent", state="disabled", wrap="word")
        self.log_box.pack(fill="both", expand=True, padx=8, pady=(2, 10))
        try:
            self.log_box.tag_config("ok",   foreground=SUCCESS)
            self.log_box.tag_config("err",  foreground=DANGER)
            self.log_box.tag_config("warn", foreground=WARNING)
        except Exception:
            pass

    # ─────────────────────────────  Results tab  ─────────────────────────────
    def _build_results_tab(self):
        tab = self.tab_results

        # Stat cards
        self.stats_frame = ctk.CTkFrame(tab, fg_color="transparent")
        self.stats_frame.pack(fill="x", padx=24, pady=(20, 10))
        self.stat_labels: dict = {}
        stat_defs = [("Bills", "🧾"), ("HSN Codes", "🔖"),
                     ("Taxable Value", "₹"), ("Total Tax", "🏛")]
        for i, (key, icon) in enumerate(stat_defs):
            card = ctk.CTkFrame(self.stats_frame, corner_radius=12, fg_color=CARD_BG, height=84)
            card.pack(side="left", fill="both", expand=True,
                      padx=(0 if i == 0 else 6, 0 if i == len(stat_defs)-1 else 6))
            card.pack_propagate(False)
            ctk.CTkLabel(card, text=f"{icon}  {key}", text_color=MUTED,
                         font=self._font(11), anchor="w").pack(fill="x", padx=16, pady=(14, 0))
            v = ctk.CTkLabel(card, text="—", font=self._font(22, "bold"), anchor="w")
            v.pack(fill="x", padx=16)
            self.stat_labels[key] = v

        # Unknown-HSN warning banner
        self.flag_frame = ctk.CTkFrame(tab, fg_color=("#FFF3CD", "#3D2F00"), corner_radius=10)
        self.flag_label = ctk.CTkLabel(self.flag_frame, text="", wraplength=720,
                                        text_color=("#7D5A00", "#FFD060"),
                                        font=self._font(11), anchor="w")
        self.flag_label.pack(side="left", padx=14, pady=10, fill="x", expand=True)
        self.flag_btn = ctk.CTkButton(self.flag_frame, text="Enter HSN →", width=120,
                                       fg_color=WARNING, hover_color="#D98F1B",
                                       text_color="black", font=self._font(11, "bold"),
                                       command=self._open_flag_dialog)
        self.flag_btn.pack(side="right", padx=14, pady=10)

        ctk.CTkLabel(tab, text="Double-click any cell to edit  ·  All values editable before export",
                     text_color=MUTED, font=self._font(10)).pack(anchor="w", padx=26)

        # Table
        tbl_card = ctk.CTkFrame(tab, corner_radius=12, fg_color=CARD_BG)
        tbl_card.pack(fill="both", expand=True, padx=24, pady=(6, 8))
        cols = ("#", "HSN Code", "UQC", "Quantity", "Taxable Value", "IGST", "CGST", "SGST")
        vsb = ttk.Scrollbar(tbl_card, orient="vertical")
        vsb.pack(side="right", fill="y", pady=8)
        self.tree = ttk.Treeview(tbl_card, columns=cols, show="headings",
                                  yscrollcommand=vsb.set, selectmode="browse")
        vsb.config(command=self.tree.yview)
        col_widths = [44, 120, 70, 95, 140, 115, 115, 115]
        for col, w in zip(cols, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w,
                              anchor="e" if col not in ("#", "HSN Code", "UQC") else "w")
        self.tree.pack(fill="both", expand=True, padx=8, pady=8)
        self.tree.bind("<Double-1>", self._edit_cell)
        self._style_tree()

        # Action row
        btn_row = ctk.CTkFrame(tab, fg_color="transparent")
        btn_row.pack(fill="x", padx=24, pady=(0, 16))
        ctk.CTkButton(btn_row, text="⬇  Download Excel", font=self._font(13, "bold"),
                       height=42, corner_radius=10, fg_color=SUCCESS, hover_color="#27AE60",
                       command=self._download_excel).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="+ Add Row", height=42, fg_color="transparent",
                       border_width=1, text_color=MUTED, font=self._font(12),
                       command=self._add_row).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="Delete Selected", height=42, fg_color="transparent",
                       border_width=1, text_color=MUTED, font=self._font(12),
                       command=self._delete_row).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="Start Over", height=42, fg_color="transparent",
                       border_width=1, text_color=DANGER, font=self._font(12),
                       command=self._reset).pack(side="right")

    def _style_tree(self):
        dark = ctk.get_appearance_mode() == "Dark"
        bg      = "#23262E" if dark else "#FFFFFF"
        fg      = "#E8E8E8" if dark else "#1A1A1A"
        stripe  = "#2A2E38" if dark else "#F2F5FB"
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background=bg, foreground=fg, fieldbackground=bg,
                         rowheight=30, font=(FONT_FAMILY, 10), borderwidth=0)
        style.configure("Treeview.Heading", background=ACCENT, foreground="white",
                         font=(FONT_FAMILY, 10, "bold"), borderwidth=0)
        style.map("Treeview", background=[("selected", ACCENT)])
        style.map("Treeview.Heading", background=[("active", ACCENT_HOVER)])
        if hasattr(self, "tree"):
            self.tree.tag_configure("odd",  background=bg)
            self.tree.tag_configure("even", background=stripe)
            self.tree.tag_configure("unknown", foreground=WARNING)

    # ─────────────────────────────  Sidebar file list  ─────────────────────────────
    def _refresh_file_sidebar(self):
        for w in self.file_listbox_frame.winfo_children():
            w.destroy()
        colors_map = {"PDF": DANGER, "JPG": "#3498DB", "JPEG": "#3498DB",
                      "PNG": "#3498DB", "XLSX": SUCCESS, "XLS": SUCCESS, "CSV": SUCCESS}
        for f in self.files:
            ext = Path(f).suffix.upper().lstrip(".")
            c = colors_map.get(ext, "#888888")
            row = ctk.CTkFrame(self.file_listbox_frame, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=ext, width=44, fg_color=c, corner_radius=5,
                          text_color="white",
                          font=self._font(9, "bold")).pack(side="left")
            name = Path(f).name
            disp = name[:20] + "…" if len(name) > 22 else name
            ctk.CTkLabel(row, text=disp, font=self._font(10), anchor="w").pack(
                side="left", padx=6, fill="x", expand=True)
            ctk.CTkButton(row, text="✕", width=24, height=22, fg_color="transparent",
                          text_color=MUTED, hover_color=("#E0E0E0", "#3A3A3A"),
                          command=lambda p=f: self._remove_file(p)).pack(side="right")
        n = len(self.files)
        self.file_count_chip.configure(text=str(n))
        self.upload_status.configure(text=f"{n} file{'s' if n != 1 else ''} ready."
                                          if n else "No files selected.")
        self.process_btn.configure(state="normal" if n > 0 else "disabled")

    def _remove_file(self, path):
        if path in self.files:
            self.files.remove(path)
        self._refresh_file_sidebar()

    # ─────────────────────────────  Actions  ─────────────────────────────
    def _save_key(self):
        self.api_key = self.api_entry.get().strip()
        self._save_config()
        self.key_status.configure(
            text="●  Key saved" if self.api_key else "○  No key saved",
            text_color=SUCCESS if self.api_key else MUTED)

    def _browse_files(self):
        paths = filedialog.askopenfilenames(
            title="Select bills",
            filetypes=[("All supported", "*.jpg *.jpeg *.png *.pdf *.xlsx *.xls *.csv"),
                       ("Images", "*.jpg *.jpeg *.png"), ("PDF", "*.pdf"),
                       ("Excel/CSV", "*.xlsx *.xls *.csv")])
        for p in paths:
            if p not in self.files:
                self.files.append(p)
        if paths:
            self._refresh_file_sidebar()

    def _clear_files(self):
        self.files.clear()
        self._refresh_file_sidebar()

    def _start_processing(self):
        key = self.api_entry.get().strip()
        needs_ai = any(Path(f).suffix.lower() in (".jpg", ".jpeg", ".png", ".pdf")
                       for f in self.files)
        if needs_ai and not key:
            messagebox.showerror("API Key Missing",
                                  "Images/PDFs need Claude Vision.\n"
                                  "Enter your Anthropic API key (console.anthropic.com), "
                                  "or upload Excel/CSV files only.")
            return
        self.api_key = key
        if key:
            self._save_config()
        self.hsn_map = {}
        self._clear_log()
        self.process_btn.configure(state="disabled")
        self.after(0, self.progress_bar.set, 0)
        self.progress_pct.configure(text="0%")
        self.tabview.set("  2 · Process  ")
        self.after(200, self._launch_thread)

    def _launch_thread(self):
        self._log("Starting processing...\n")
        threading.Thread(target=self._process_all, daemon=True).start()

    # ─────────────────────────────  Pipeline  ─────────────────────────────
    def _process_all(self):
        import traceback
        client = None
        needs_ai = any(Path(f).suffix.lower() in (".jpg", ".jpeg", ".png", ".pdf")
                       for f in self.files)
        if needs_ai:
            try:
                client = anthropic.Anthropic(api_key=self.api_key)
                self._log("✓ Claude client ready.\n", "ok")
            except Exception as e:
                self._log(f"✗ Failed to initialise Claude client:\n  {e}", "err")
                self.after(0, lambda: self.process_btn.configure(state="normal"))
                return

        total = len(self.files)
        self._log(f"Total files: {total}\n")

        for idx, fpath in enumerate(self.files):
            name = Path(fpath).name
            ext  = Path(fpath).suffix.lower()
            self._log(f"\n── [{idx+1}/{total}] {name}")
            self._set_progress_label(f"Processing {idx+1}/{total}: {name}")
            try:
                rows: list = []
                if ext in (".jpg", ".jpeg", ".png"):
                    rows = self._process_image(client, fpath)
                elif ext == ".pdf":
                    rows = self._process_pdf(client, fpath)
                elif ext in (".xlsx", ".xls", ".csv"):
                    rows = self._process_excel(fpath)
                else:
                    self._log("  ✗ Unsupported format, skipping.", "warn")
                    continue
                self._merge_rows(rows, name)
                self._log(f"  ✓ {len(rows)} line item(s) extracted.", "ok")
            except Exception as e:
                self._log(f"  ✗ ERROR: {e}", "err")
                self._log(traceback.format_exc(), "err")

            pct = (idx + 1) / total
            self.after(0, lambda v=pct: self.progress_bar.set(v))
            self.after(0, lambda v=pct: self.progress_pct.configure(text=f"{int(v*100)}%"))

        self._log("\n── All files processed. ──────────────────", "ok")
        self.after(0, self._show_results)

    def _process_image(self, client, fpath: str) -> list:
        self._log("  → Reading image via Claude Haiku Vision...")
        with open(fpath, "rb") as f:
            b64 = base64.standard_b64encode(f.read()).decode()
        ext = Path(fpath).suffix.lower()
        mt = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
        return self._call_claude(client, b64, mt, Path(fpath).name)

    def _process_pdf(self, client, fpath: str) -> list:
        self._log("  → Converting PDF pages to images...")
        doc = fitz.open(fpath)
        all_rows: list = []
        for i in range(len(doc)):
            page = doc[i]
            self._log(f"     Page {i+1}/{len(doc)}...")
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
            b64 = base64.standard_b64encode(pix.tobytes("jpeg")).decode()
            rows = self._call_claude(client, b64, "image/jpeg", f"{Path(fpath).name} p{i+1}")
            all_rows.extend(rows)
        doc.close()
        return all_rows

    def _call_claude(self, client, b64: str, media_type: str, label: str) -> list:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2000,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                {"type": "text", "text": EXTRACT_PROMPT}
            ]}]
        )
        raw = "".join(b.text for b in msg.content if hasattr(b, "text"))
        raw = re.sub(r"```json|```", "", raw).strip()
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            m = re.search(r"\[.*\]", raw, re.DOTALL)
            if not m:
                raise ValueError(f"No JSON array in response for {label}")
            parsed = json.loads(m.group(0))
        self._log(f"     AI found {len(parsed)} item(s) in {label}")
        return parsed

    def _process_excel(self, fpath: str) -> list:
        self._log("  → Parsing Excel/CSV (no AI, free)...")
        ext = Path(fpath).suffix.lower()
        if ext == ".csv":
            with open(fpath, newline="", encoding="utf-8-sig") as f:
                raw = list(csv.reader(f))
        else:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            ws = wb.active
            if ws is None:
                raise ValueError("No active sheet found")
            raw = [[cell.value for cell in row] for row in ws.iter_rows()]

        if len(raw) < 2:
            raise ValueError("Sheet appears empty")

        header_idx = 0
        for i, row in enumerate(raw[:10]):
            joined = " ".join(str(c or "").lower().strip() for c in row)
            if any(a in joined for a in HSN_ALIASES + QTY_ALIASES):
                header_idx = i
                break

        headers = [str(c or "") for c in raw[header_idx]]
        self._log(f"     Headers: {[h for h in headers if h]}")

        hsn_i  = match_col(headers, HSN_ALIASES)
        qty_i  = match_col(headers, QTY_ALIASES)
        uqc_i  = match_col(headers, UQC_ALIASES)
        tax_i  = match_col(headers, TAX_ALIASES)
        igst_i = match_col(headers, IGST_ALIASES)
        cgst_i = match_col(headers, CGST_ALIASES)
        sgst_i = match_col(headers, SGST_ALIASES)

        rows = []
        for row in raw[header_idx + 1:]:
            if all(not str(c or "").strip() for c in row):
                continue
            hsn = str(row[hsn_i] or "").strip() if hsn_i >= 0 else "UNKNOWN"
            rows.append({
                "hsn":           hsn or "UNKNOWN",
                "uqc":           str(row[uqc_i] or "NOS").strip() if uqc_i >= 0 else "NOS",
                "quantity":      to_num(row[qty_i])  if qty_i  >= 0 else 0,
                "taxable_value": to_num(row[tax_i])  if tax_i  >= 0 else 0,
                "igst":          to_num(row[igst_i]) if igst_i >= 0 else 0,
                "cgst":          to_num(row[cgst_i]) if cgst_i >= 0 else 0,
                "sgst":          to_num(row[sgst_i]) if sgst_i >= 0 else 0,
            })
        self._log(f"     Parsed {len(rows)} data rows.")
        return rows

    def _merge_rows(self, rows: list, source: str):
        for r in rows:
            key = str(r.get("hsn") or "UNKNOWN").strip() or "UNKNOWN"
            if key not in self.hsn_map:
                self.hsn_map[key] = {"hsn": key, "uqc": r.get("uqc", "NOS"),
                                      "quantity": 0.0, "taxable_value": 0.0,
                                      "igst": 0.0, "cgst": 0.0, "sgst": 0.0}
            m = self.hsn_map[key]
            m["quantity"]      += to_num(r.get("quantity", 0))
            m["taxable_value"] += to_num(r.get("taxable_value", 0))
            m["igst"]          += to_num(r.get("igst", 0))
            m["cgst"]          += to_num(r.get("cgst", 0))
            m["sgst"]          += to_num(r.get("sgst", 0))

    # ─────────────────────────────  Results  ─────────────────────────────
    def _show_results(self):
        self._set_progress_label("Done!")
        self.tabview.set("  3 · Review & Export  ")
        self._refresh_results()

    def _refresh_results(self):
        entries = list(self.hsn_map.values())
        known   = [e for e in entries if e["hsn"] != "UNKNOWN"]
        tv      = sum(e["taxable_value"] for e in entries)
        tax     = sum(e["igst"] + e["cgst"] + e["sgst"] for e in entries)
        self.stat_labels["Bills"].configure(text=str(len(self.files)))
        self.stat_labels["HSN Codes"].configure(text=str(len(known)))
        self.stat_labels["Taxable Value"].configure(text=f"₹{tv:,.0f}")
        self.stat_labels["Total Tax"].configure(text=f"₹{tax:,.0f}")

        if "UNKNOWN" in self.hsn_map:
            u = self.hsn_map["UNKNOWN"]
            self.flag_frame.pack(fill="x", padx=24, pady=(0, 6))
            self.flag_label.configure(
                text=f"⚠  Some items missing HSN codes "
                     f"(Qty: {u['quantity']:.2f} | Taxable: ₹{u['taxable_value']:,.2f}). "
                     f"Click 'Enter HSN' to assign them.")
        else:
            self.flag_frame.pack_forget()

        for row in self.tree.get_children():
            self.tree.delete(row)
        for i, e in enumerate(entries, 1):
            tags = ["unknown"] if e["hsn"] == "UNKNOWN" else ["even" if i % 2 == 0 else "odd"]
            self.tree.insert("", "end", iid=str(i), tags=tags,
                              values=(i, e["hsn"], e["uqc"],
                                      f"{e['quantity']:.2f}", f"{e['taxable_value']:.2f}",
                                      f"{e['igst']:.2f}", f"{e['cgst']:.2f}", f"{e['sgst']:.2f}"))
        self.process_btn.configure(state="normal")

    def _edit_cell(self, event):
        item = self.tree.focus()
        if not item:
            return
        col_num = int(self.tree.identify_column(event.x).replace("#", "")) - 1
        cols = ("#", "HSN Code", "UQC", "Quantity", "Taxable Value", "IGST", "CGST", "SGST")
        if col_num == 0:
            return
        col_name     = cols[col_num]
        current_vals = self.tree.item(item, "values")
        old_hsn      = current_vals[1]
        dialog  = ctk.CTkInputDialog(text=f"Edit {col_name}:", title=f"Edit {col_name}")
        new_val = dialog.get_input()
        if new_val is None:
            return
        field_map = {"HSN Code": "hsn", "UQC": "uqc", "Quantity": "quantity",
                     "Taxable Value": "taxable_value", "IGST": "igst",
                     "CGST": "cgst", "SGST": "sgst"}
        field = field_map.get(col_name)
        if field and old_hsn in self.hsn_map:
            entry = self.hsn_map[old_hsn]
            if field == "hsn":
                del self.hsn_map[old_hsn]
                entry["hsn"] = new_val
                self.hsn_map[new_val] = entry
            else:
                try:
                    entry[field] = float(new_val)
                except ValueError:
                    entry[field] = new_val
        self._refresh_results()

    def _add_row(self):
        key = f"NEW_{len(self.hsn_map)}"
        self.hsn_map[key] = {"hsn": key, "uqc": "NOS", "quantity": 0.0,
                              "taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0}
        self._refresh_results()

    def _delete_row(self):
        item = self.tree.focus()
        if not item:
            return
        hsn = self.tree.item(item, "values")[1]
        if hsn in self.hsn_map:
            del self.hsn_map[hsn]
        self._refresh_results()

    def _open_flag_dialog(self):
        if "UNKNOWN" not in self.hsn_map:
            return
        u = self.hsn_map["UNKNOWN"]
        win = ctk.CTkToplevel(self)
        win.title("Assign HSN to unidentified items")
        win.geometry("440x300")
        win.grab_set()
        ctk.CTkLabel(win, text="Items without HSN code found:",
                     font=self._font(13, "bold")).pack(pady=(20, 4), padx=20, anchor="w")
        ctk.CTkLabel(win, wraplength=390, text_color=MUTED, font=self._font(11),
                     text=f"Qty: {u['quantity']:.2f}  |  Taxable: ₹{u['taxable_value']:,.2f}  |  "
                          f"IGST: ₹{u['igst']:.2f}  |  CGST: ₹{u['cgst']:.2f}  |  SGST: ₹{u['sgst']:.2f}"
                     ).pack(padx=20, anchor="w")
        ctk.CTkLabel(win, text="HSN Code:", anchor="w",
                     font=self._font(11)).pack(fill="x", padx=20, pady=(16, 4))
        hsn_entry = ctk.CTkEntry(win, placeholder_text="e.g. 998314")
        hsn_entry.pack(fill="x", padx=20)
        ctk.CTkLabel(win, text="UQC:", anchor="w",
                     font=self._font(11)).pack(fill="x", padx=20, pady=(8, 4))
        uqc_entry = ctk.CTkEntry(win, placeholder_text="e.g. NOS")
        uqc_entry.insert(0, u.get("uqc", "NOS"))
        uqc_entry.pack(fill="x", padx=20)
        def apply():
            hsn = hsn_entry.get().strip()
            uqc = uqc_entry.get().strip() or "NOS"
            if not hsn:
                messagebox.showwarning("Required", "Please enter an HSN code.", parent=win)
                return
            if hsn in self.hsn_map:
                for k in ["quantity", "taxable_value", "igst", "cgst", "sgst"]:
                    self.hsn_map[hsn][k] += u[k]
            else:
                self.hsn_map[hsn] = {**u, "hsn": hsn, "uqc": uqc}
            del self.hsn_map["UNKNOWN"]
            win.destroy()
            self._refresh_results()
        ctk.CTkButton(win, text="Apply HSN Code", fg_color=ACCENT,
                      hover_color=ACCENT_HOVER, command=apply).pack(pady=16)

    # ─────────────────────────────  Excel export  ─────────────────────────────
    def _download_excel(self):
        if not self.hsn_map:
            messagebox.showwarning("No Data", "Nothing to export yet.")
            return
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile="GSTR1_HSN_Summary.xlsx")
        if not save_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "HSN Summary"  # type: ignore[union-attr]

        hfill  = PatternFill("solid", start_color="1F4E79")
        hfont  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        border = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
                        top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"))
        center = Alignment(horizontal="center", vertical="center")
        right  = Alignment(horizontal="right",  vertical="center")
        alt1   = PatternFill("solid", start_color="EBF3FB")
        alt2   = PatternFill("solid", start_color="FFFFFF")

        headers = ["HSN/SAC Code", "UQC", "Total Quantity", "Total Taxable Value",
                   "Integrated Tax (IGST)", "Central Tax (CGST)", "State/UT Tax (SGST)"]
        col_w   = [16, 8, 14, 18, 20, 18, 18]

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)  # type: ignore[union-attr]
            c.fill = hfill; c.font = hfont; c.border = border; c.alignment = center
            ws.column_dimensions[get_column_letter(ci)].width = col_w[ci-1]  # type: ignore[union-attr]
        ws.row_dimensions[1].height = 20  # type: ignore[union-attr]

        entries = list(self.hsn_map.values())
        for ri, e in enumerate(entries, 2):
            fill = alt1 if ri % 2 == 0 else alt2
            row_data = [e["hsn"], e["uqc"], round(e["quantity"], 2), round(e["taxable_value"], 2),
                        round(e["igst"], 2), round(e["cgst"], 2), round(e["sgst"], 2)]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)  # type: ignore[union-attr]
                c.fill = fill; c.border = border; c.font = Font(name="Arial", size=10)
                c.alignment = right if ci > 2 else Alignment(vertical="center")
                if ci > 2:
                    c.number_format = "#,##0.00"

        tr = len(entries) + 2
        for ci in range(3, 8):
            col = get_column_letter(ci)
            c = ws.cell(row=tr, column=ci)  # type: ignore[union-attr]
            c.value = f"=SUM({col}2:{col}{tr-1})"
            c.font = Font(bold=True, name="Arial", size=10)
            c.fill = PatternFill("solid", start_color="D6E4F0")
            c.border = border; c.alignment = right; c.number_format = "#,##0.00"

        ws.cell(row=tr, column=2).value = "TOTAL"  # type: ignore[union-attr]
        ws.cell(row=tr, column=2).font  = Font(bold=True, name="Arial", size=10)  # type: ignore[union-attr]
        ws.cell(row=tr, column=2).fill  = PatternFill("solid", start_color="D6E4F0")  # type: ignore[union-attr]
        ws.cell(row=tr, column=2).border = border  # type: ignore[union-attr]
        ws.freeze_panes = "A2"  # type: ignore[union-attr]
        wb.save(save_path)
        messagebox.showinfo("Exported", f"Saved to:\n{save_path}")

    # ─────────────────────────────  Utilities  ─────────────────────────────
    def _log(self, msg: str, tag: str = ""):
        def _do():
            self.log_box.configure(state="normal")
            if tag:
                try:
                    self.log_box.insert("end", msg + "\n", tag)
                except Exception:
                    self.log_box.insert("end", msg + "\n")
            else:
                self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _do)

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _set_progress_label(self, text: str):
        self.after(0, lambda t=text: self.progress_label.configure(text=t))

    def _reset(self):
        self.files.clear()
        self.hsn_map.clear()
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._refresh_file_sidebar()
        self._clear_log()
        self.after(0, self.progress_bar.set, 0)
        self.progress_pct.configure(text="0%")
        self.progress_label.configure(text="Waiting...")
        self.flag_frame.pack_forget()
        for k in self.stat_labels:
            self.stat_labels[k].configure(text="—")
        self.tabview.set("  1 · Upload  ")


if __name__ == "__main__":
    app = App()
    app.mainloop()
